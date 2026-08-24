import hashlib
import io
import math
import zipfile
from typing import Any

import numpy as np
import pandas as pd
from scipy import stats

MAX_ROWS = 100_000
MAX_COLUMNS = 500
MAX_CELL_LENGTH = 32_000
MAX_XLSX_ENTRIES = 5_000
MAX_XLSX_UNCOMPRESSED = 250 * 1024 * 1024
MAX_XLSX_SHEETS = 50
ENGINE_VERSION = "baseerah-stats-1.0"

def safe_csv_value(value: Any) -> Any:
    """Neutralize spreadsheet formulas without changing stored research data."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def load_tabular(content: bytes, filename: str) -> pd.DataFrame:
    lower = filename.lower()
    if lower.endswith(".csv"):
        frame = pd.read_csv(io.BytesIO(content))
    elif lower.endswith(".xlsx"):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                entries=archive.infolist()
                if len(entries)>MAX_XLSX_ENTRIES: raise ValueError("XLSX exceeds the ZIP entry safety limit")
                total=0
                for entry in entries:
                    normalized=entry.filename.replace("\\","/")
                    if normalized.startswith("/") or ".." in normalized.split("/") or ":" in normalized: raise ValueError("XLSX contains an unsafe ZIP path")
                    total += entry.file_size
                    if total>MAX_XLSX_UNCOMPRESSED: raise ValueError("XLSX exceeds the uncompressed-size safety limit")
        except zipfile.BadZipFile as exc: raise ValueError("Malformed XLSX container") from exc
        from openpyxl import load_workbook
        workbook=load_workbook(io.BytesIO(content),read_only=True,data_only=True)
        try:
            if len(workbook.sheetnames)>MAX_XLSX_SHEETS: raise ValueError("XLSX exceeds the worksheet safety limit")
        finally: workbook.close()
        frame = pd.read_excel(io.BytesIO(content), engine="openpyxl")
    else:
        raise ValueError("Only CSV and XLSX datasets are supported")
    if frame.shape[0] > MAX_ROWS or frame.shape[1] > MAX_COLUMNS:
        raise ValueError(f"Dataset exceeds safety limits ({MAX_ROWS} rows, {MAX_COLUMNS} columns)")
    if any(frame[col].astype("string").str.len().max() > MAX_CELL_LENGTH for col in frame.columns):
        raise ValueError("Dataset contains a cell longer than the safety limit")
    frame.columns = [str(col).strip() for col in frame.columns]
    if not all(frame.columns) or len(set(frame.columns)) != len(frame.columns):
        raise ValueError("Column names must be non-empty and unique")
    return frame


def frame_records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    clean = frame.replace({np.nan: None, np.inf: None, -np.inf: None})
    return clean.to_dict(orient="records")


def fingerprint(records: list[dict[str, Any]]) -> str:
    import json
    canonical = json.dumps(records, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def infer_variables(frame: pd.DataFrame) -> list[dict[str, Any]]:
    result = []
    for name in frame.columns:
        series = frame[name]
        if pd.api.types.is_bool_dtype(series):
            data_type, level = "BOOLEAN", "BINARY"
        elif pd.api.types.is_integer_dtype(series):
            data_type, level = "INTEGER", "RATIO"
        elif pd.api.types.is_numeric_dtype(series):
            data_type, level = "FLOAT", "RATIO"
        elif pd.api.types.is_datetime64_any_dtype(series):
            data_type, level = "DATETIME", "DATE_TIME"
        else:
            unique = series.nunique(dropna=True)
            data_type = "CATEGORY" if unique <= min(30, max(2, len(series) // 4)) else "STRING"
            level = "NOMINAL" if data_type == "CATEGORY" else "FREE_TEXT"
        normalized = str(name).lower().replace("_", " ")
        identifier = any(token in normalized for token in ("name", "email", "phone", "student id", "national id", "الاسم", "الهاتف", "الهوية"))
        result.append({"name": str(name), "data_type": data_type, "measurement_level": level,
                       "role": "IDENTIFIER" if identifier else "OTHER", "sensitive": identifier,
                       "identifier": identifier})
    return result


def quality_scan(frame: pd.DataFrame) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    total_cells = max(1, frame.shape[0] * frame.shape[1])
    missing = int(frame.isna().sum().sum())
    duplicate_rows = int(frame.duplicated().sum())
    for column in frame.columns:
        count = int(frame[column].isna().sum())
        if count:
            issues.append({"variable_name": str(column), "issue_type": "MISSING_VALUES",
                           "severity": "HIGH" if count / max(1, len(frame)) >= .2 else "MEDIUM",
                           "details": {"count": count, "percentage": round(count / max(1, len(frame)) * 100, 2)}})
        if pd.api.types.is_numeric_dtype(frame[column]):
            values = frame[column].dropna().astype(float)
            if len(values) >= 4:
                q1, q3 = values.quantile([.25, .75]); iqr = q3 - q1
                outliers = int(((values < q1 - 1.5 * iqr) | (values > q3 + 1.5 * iqr)).sum()) if iqr else 0
                if outliers:
                    issues.append({"variable_name": str(column), "issue_type": "OUTLIER_IQR", "severity": "MEDIUM",
                                   "details": {"count": outliers, "method": "IQR_1.5"}})
    if duplicate_rows:
        issues.append({"variable_name": None, "issue_type": "EXACT_DUPLICATE", "severity": "HIGH",
                       "details": {"count": duplicate_rows}})
    completeness = round((1 - missing / total_cells) * 100, 2)
    penalty = min(35, duplicate_rows * 2 + sum(4 if i["severity"] == "HIGH" else 2 for i in issues))
    summary = {"completeness": completeness, "validity": 100.0, "consistency": max(0, 100 - penalty),
               "missing_values": missing, "duplicates": duplicate_rows,
               "outliers": sum(i["details"]["count"] for i in issues if i["issue_type"] == "OUTLIER_IQR"),
               "quality_score": round(max(0, completeness * .6 + (100 - penalty) * .4), 2)}
    return summary, issues


def decide_test(objective: str, dv_level: str, groups: int, paired: bool, normal: bool) -> dict[str, str]:
    objective = objective.upper(); level = dv_level.upper()
    if objective == "ASSOCIATION" and level in {"NOMINAL", "BINARY"}:
        return {"test": "CHI_SQUARE", "reason": "Association between categorical variables"}
    if objective == "RELATIONSHIP" and level in {"INTERVAL", "RATIO"}:
        return {"test": "PEARSON" if normal else "SPEARMAN", "reason": "Relationship between numeric variables"}
    if groups == 2:
        if paired:
            return {"test": "PAIRED_T_TEST" if normal else "WILCOXON", "reason": "Two related measurements"}
        return {"test": "INDEPENDENT_T_TEST" if normal else "MANN_WHITNEY", "reason": "Two independent groups"}
    if groups > 2:
        return {"test": "ONE_WAY_ANOVA" if normal else "KRUSKAL_WALLIS", "reason": "Comparison across multiple groups"}
    raise ValueError("Insufficient design information for a valid recommendation")


def run_analysis(records: list[dict[str, Any]], kind: str, config: dict[str, Any]) -> dict[str, Any]:
    frame = pd.DataFrame(records); kind = kind.upper(); alpha = float(config.get("alpha", .05))
    if not 0 < alpha < 1: raise ValueError("alpha must be between 0 and 1")
    if kind == "DESCRIPTIVES":
        cols = config.get("variables") or list(frame.select_dtypes(include="number").columns)
        output = {}
        for col in cols:
            values = pd.to_numeric(frame[col], errors="coerce").dropna()
            if len(values) == 0: raise ValueError(f"Variable {col} contains no numeric observations")
            output[col] = {"n": int(len(values)), "missing": int(frame[col].isna().sum()), "mean": float(values.mean()),
                           "sd": float(values.std(ddof=1)) if len(values) > 1 else None, "variance": float(values.var(ddof=1)) if len(values) > 1 else None, "median": float(values.median()),
                           "min": float(values.min()), "max": float(values.max())}
        return {"analysis": kind, "method": "Sample descriptives (SD/variance use ddof=1)", "estimates": output, "warnings": []}
    if kind == "INDEPENDENT_T_TEST":
        outcome, group = config["outcome"], config["group"]
        groups = [(name, pd.to_numeric(part[outcome], errors="coerce").dropna()) for name, part in frame.groupby(group)]
        if len(groups) != 2 or any(len(v) < 2 for _, v in groups): raise ValueError("Independent t-test requires exactly two groups with at least two observations each")
        (n1, a), (n2, b) = groups; test = stats.ttest_ind(a, b, equal_var=False); diff = float(a.mean() - b.mean())
        if a.var(ddof=1)==0 and b.var(ddof=1)==0: raise ValueError("Welch t-test is undefined when both groups are constant")
        se = math.sqrt(a.var(ddof=1)/len(a) + b.var(ddof=1)/len(b)); df_num = (a.var(ddof=1)/len(a) + b.var(ddof=1)/len(b))**2
        df_den = (a.var(ddof=1)/len(a))**2/(len(a)-1) + (b.var(ddof=1)/len(b))**2/(len(b)-1); df = df_num/df_den
        critical = stats.t.ppf(1-alpha/2, df); pooled = math.sqrt(((len(a)-1)*a.var(ddof=1)+(len(b)-1)*b.var(ddof=1))/(len(a)+len(b)-2))
        values_to_check=[test.statistic,test.pvalue,df,diff,se,pooled]
        if not all(np.isfinite(float(value)) for value in values_to_check): raise ValueError("Welch t-test is undefined for constant or non-finite input")
        return {"analysis": kind, "method": "Welch independent samples t-test", "groups": [str(n1), str(n2)], "n": [len(a), len(b)], "means": [float(a.mean()), float(b.mean())],
                "mean_difference": diff, "confidence_interval": [float(diff-critical*se), float(diff+critical*se)], "confidence_level": 1-alpha,
                "statistic": float(test.statistic), "df": float(df), "p_value": float(test.pvalue), "effect_size": {"name": "Cohen_d_pooled_sd", "value": diff/pooled}, "warnings": []}
    if kind in {"PEARSON", "SPEARMAN"}:
        x, y = config["x"], config["y"]; pair = frame[[x,y]].apply(pd.to_numeric, errors="coerce").dropna()
        if len(pair) < 3: raise ValueError("Correlation requires at least three complete pairs")
        if pair[x].nunique()<2 or pair[y].nunique()<2: raise ValueError("Correlation is undefined for a constant variable")
        res = stats.pearsonr(pair[x], pair[y]) if kind == "PEARSON" else stats.spearmanr(pair[x], pair[y])
        if not np.isfinite(float(res.statistic)) or not np.isfinite(float(res.pvalue)): raise ValueError("Correlation produced a non-finite result")
        return {"analysis": kind, "method": "Pearson product-moment correlation" if kind=="PEARSON" else "Spearman rank correlation with tie correction", "n": len(pair), "statistic": float(res.statistic), "p_value": float(res.pvalue), "warnings": []}
    if kind == "CHI_SQUARE":
        table = pd.crosstab(frame[config["x"]], frame[config["y"]]);
        if table.shape[0] < 2 or table.shape[1] < 2: raise ValueError("Chi-square requires a table of at least 2x2")
        chi, p, df, expected = stats.chi2_contingency(table); n = int(table.values.sum()); denom = min(table.shape)-1
        if not all(np.isfinite(float(value)) for value in (chi,p,expected.min())): raise ValueError("Chi-square produced a non-finite result")
        warnings=["Expected cell frequency below 5; chi-square approximation may be unreliable"] if expected.min()<5 else []
        return {"analysis": kind, "method": "Pearson chi-square test of independence", "n": n, "statistic": float(chi), "df": int(df), "p_value": float(p),
                "effect_size": {"name": "Cramers_V", "value": math.sqrt(chi/(n*denom))}, "minimum_expected": float(expected.min()), "expected_frequencies": expected.tolist(), "warnings": warnings}
    raise ValueError(f"Analysis type {kind} is not implemented")
